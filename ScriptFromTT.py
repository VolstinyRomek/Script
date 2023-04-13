import openpyxl
import requests
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

workbook = openpyxl.load_workbook('TIKTOKLINKS.xlsx')
worksheet = workbook.active

driver = webdriver.Chrome(ChromeDriverManager().install())
driver.get('https://ssstik.io/en')

for i in range(1, worksheet.max_row + 1):
    link = worksheet.cell(row=i, column=1).value

    input_box = driver.find_element_by_xpath("//input[@name='url']")
    input_box.send_keys(link)

    button = driver.find_element_by_xpath("//button[@type='submit']")
    button.click()

    try:
        popup = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//div[@class='popup']"))
        )
        driver.close()
    except:
        pass

    download_button = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, "//a[@class='Without watermark-button']"))
    )
    download_link = download_button.get_attribute('href')

    response = requests.get(download_link)
    with open(f'файл_{i}.zip', 'wb') as file:
        file.write(response.content)

    input_box.clear()

driver.quit()
workbook.close()
